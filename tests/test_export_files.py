"""Tests for the CSV / XLSX export writers."""

from __future__ import annotations

from datetime import datetime, time

from openpyxl import load_workbook

from custom_components.smgw_han.aggregation import build_daily_summary
from custom_components.smgw_han.const import OBIS_EXPORT, OBIS_IMPORT
from custom_components.smgw_han.export_files import (
    _pivot_by_timestamp,
    write_readings_csv,
    write_xlsx,
)
from custom_components.smgw_han.smgw_client import MeterReading


def _imp(ts: datetime, value: float) -> MeterReading:
    return MeterReading(ts, OBIS_IMPORT, value, "kWh", "valid")


def _exp(ts: datetime, value: float) -> MeterReading:
    return MeterReading(ts, OBIS_EXPORT, value, "kWh", "valid")


# A 05:00 import reading without a matching export reading exercises the
# "missing OBIS code -> empty cell, never a fabricated zero" behaviour.
READINGS = [
    _imp(datetime(2026, 5, 15, 0, 0, 1), 1000.0),
    _exp(datetime(2026, 5, 15, 0, 0, 1), 500.0),
    _imp(datetime(2026, 5, 15, 5, 0, 1), 1002.5),
    _imp(datetime(2026, 5, 16, 0, 0, 1), 1010.0),
    _exp(datetime(2026, 5, 16, 0, 0, 1), 503.0),
]


def test_pivot_leaves_missing_export_as_none():
    rows = {ts: (imp, exp, q) for ts, imp, exp, q in _pivot_by_timestamp(READINGS)}
    imp, exp, _q = rows[datetime(2026, 5, 15, 5, 0, 1)]
    assert imp == 1002.5
    assert exp is None  # -> rendered as an empty cell, not 0


def test_write_csv_wide_format(tmp_path):
    path = tmp_path / "out.csv"
    write_readings_csv(path, READINGS)
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    assert lines[0] == (
        "Zeitstempel;1.8.0 Bezug (kWh);2.8.0 Einspeisung (kWh);Qualität"
    )
    row = next(line for line in lines if line.startswith("2026-05-15 05:00:01"))
    # Columns: timestamp;import;export;quality -> export must be blank.
    assert row.split(";")[2] == ""


GO_ZONES = [(time(0, 0), "Go"), (time(5, 0), "Standard")]


def test_write_xlsx_sheets_and_rowcount(tmp_path):
    path = tmp_path / "out.xlsx"
    summary = build_daily_summary(READINGS, GO_ZONES)
    meta = {
        "meter_id": "1lgz0072999211",
        "from": "2026-05-15 00:00:00",
        "to": "2026-05-16 00:15:00",
        "zones": [("00:00", "Go"), ("05:00", "Standard")],
    }
    write_xlsx(path, READINGS, summary, meta)

    wb = load_workbook(path)
    assert set(wb.sheetnames) == {
        "Rohdaten",
        "Tagesendwerte",
        "Tarifzonen",
        "Definition",
    }
    # Rohdaten is the long dump: one header row + one row per reading.
    assert wb["Rohdaten"].max_row == 1 + len(READINGS)
    # Tagesendwerte: header + one row per summarized day.
    assert wb["Tagesendwerte"].max_row == 1 + len(summary)
    # Tarifzonen: one boundary column per inner switch time and one
    # consumption column per distinct zone name.
    header = [c.value for c in wb["Tarifzonen"][1]]
    assert header == [
        "Datum",
        "Bezug 00:00 (kWh)",
        "Bezug 05:00 (kWh)",
        "Bezug Folgetag 00:00 (kWh)",
        "Verbrauch Go (kWh)",
        "Verbrauch Standard (kWh)",
        "Gesamtverbrauch 00:00-24:00 (kWh)",
        "Einspeisung gesamt (kWh)",
    ]
    row = [c.value for c in wb["Tarifzonen"][2]]
    assert row == [
        "2026-05-15", 1000.0, 1002.5, 1010.0, 2.5, 7.5, 10.0, 3.0,
    ]


def test_formula_injection_is_neutralized(tmp_path):
    # User-defined zone names and gateway-supplied text must not become
    # spreadsheet formulas (leading '=', '+', '-', '@').
    evil_readings = [
        MeterReading(
            datetime(2026, 5, 15, 0, 0, 1), OBIS_IMPORT, 1000.0, "kWh", "=2+2"
        ),
        MeterReading(
            datetime(2026, 5, 16, 0, 0, 1), OBIS_IMPORT, 1010.0, "kWh", "valid"
        ),
    ]

    csv_path = tmp_path / "out.csv"
    write_readings_csv(csv_path, evil_readings)
    row = next(
        line
        for line in csv_path.read_text(encoding="utf-8-sig").splitlines()
        if line.startswith("2026-05-15 00:00:01")
    )
    assert row.split(";")[3] == "'=2+2"  # apostrophe forces text

    zones = [(time(0, 0), "=EVIL"), (time(5, 0), "Standard")]
    summary = build_daily_summary(evil_readings, zones)
    meta = {
        "meter_id": "m",
        "from": "x",
        "to": "y",
        "zones": [("00:00", "=EVIL"), ("05:00", "Standard")],
    }
    xlsx_path = tmp_path / "out.xlsx"
    write_xlsx(xlsx_path, evil_readings, summary, meta)

    wb = load_workbook(xlsx_path)
    # Rohdaten: gateway quality text neutralized, no formula cells.
    qualities = [row[4].value for row in wb["Rohdaten"].iter_rows(min_row=2)]
    assert "'=2+2" in qualities
    # Definition: zone-name-led line neutralized.
    definition_texts = [
        c.value for row in wb["Definition"].iter_rows() for c in row if c.value
    ]
    assert any(str(v).startswith("'=EVIL:") for v in definition_texts)
    assert not any(
        str(v).startswith("=") for v in definition_texts
    )


def test_write_xlsx_multi_window_zone_definition(tmp_path):
    # A zone spanning several windows (Heat-style) must be listed with all
    # its windows on the Definition sheet and get exactly one consumption
    # column on the Tarifzonen sheet.
    zones = [
        (time(0, 0), "Standard"),
        (time(2, 0), "Niedrig"),
        (time(6, 0), "Standard"),
    ]
    summary = build_daily_summary(READINGS, zones)
    meta = {
        "meter_id": "m",
        "from": "x",
        "to": "y",
        "zones": [("00:00", "Standard"), ("02:00", "Niedrig"), ("06:00", "Standard")],
    }
    path = tmp_path / "out.xlsx"
    write_xlsx(path, READINGS, summary, meta)

    wb = load_workbook(path)
    header = [c.value for c in wb["Tarifzonen"][1]]
    assert header.count("Verbrauch Standard (kWh)") == 1
    assert "Bezug 02:00 (kWh)" in header
    assert "Bezug 06:00 (kWh)" in header
    definition_texts = [
        c.value for row in wb["Definition"].iter_rows() for c in row if c.value
    ]
    assert "Standard: 00:00-02:00 + 06:00-24:00" in definition_texts
    assert "Niedrig: 02:00-06:00" in definition_texts
