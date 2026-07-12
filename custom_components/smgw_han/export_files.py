"""Synchronous CSV / XLSX writers for the export service.

Every function here performs blocking file I/O (and, for XLSX, imports
openpyxl), so they MUST be called from an executor thread via
``hass.async_add_executor_job`` — never directly on the event loop.

- CSV  -> the raw 15-minute reading dump (semicolon-separated, UTF-8 BOM,
  opens cleanly in German Excel).
- XLSX -> a multi-sheet workbook (raw data + daily end values + tariff zones
  + definitions), mirroring the standalone ``smgw_tagesendwerte_to_excel``
  layout the owner already uses.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from .aggregation import DailySummary
from .const import OBIS_EXPORT, OBIS_IMPORT
from .smgw_client import MeterReading

# Wide-format column headers for the raw-dump CSV.
RAW_HEADERS = [
    "Zeitstempel",
    "1.8.0 Bezug (kWh)",
    "2.8.0 Einspeisung (kWh)",
    "Qualität",
]


def _fmt_dt(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""


def _pivot_by_timestamp(
    readings: list[MeterReading],
) -> list[tuple[datetime, float | None, float | None, str]]:
    """Pivot long readings into one row per timestamp.

    Returns ``(timestamp, import_value, export_value, quality)`` tuples sorted
    by timestamp. A missing OBIS code for a timestamp yields ``None`` (an empty
    cell) instead of a fabricated zero — so meters without feed-in simply leave
    the 2.8.0 column blank.
    """
    rows: dict[datetime, dict[str, object]] = {}
    for r in readings:
        row = rows.setdefault(
            r.timestamp, {"import": None, "export": None, "quality": None}
        )
        if r.obis_code == OBIS_IMPORT:
            row["import"] = r.value
            row["quality"] = r.quality
        elif r.obis_code == OBIS_EXPORT:
            row["export"] = r.value
            if row["quality"] is None:
                row["quality"] = r.quality
    return [
        (ts, rows[ts]["import"], rows[ts]["export"], rows[ts]["quality"] or "")
        for ts in sorted(rows)
    ]


def write_readings_csv(path: Path, readings: list[MeterReading]) -> None:
    """Write the raw reading dump as a semicolon CSV (Excel-friendly).

    Wide format: one row per timestamp with separate 1.8.0 / 2.8.0 columns.
    """
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(RAW_HEADERS)
        for ts, imp, exp, quality in _pivot_by_timestamp(readings):
            writer.writerow(
                [
                    _fmt_dt(ts),
                    f"{imp:.4f}" if imp is not None else "",
                    f"{exp:.4f}" if exp is not None else "",
                    quality,
                ]
            )


def write_xlsx(
    path: Path,
    readings: list[MeterReading],
    daily_summary: list[DailySummary],
    meta: dict[str, Any],
) -> None:
    """Write a multi-sheet workbook. Imports openpyxl lazily.

    ``meta["zones"]`` is the tariff-zone definition as ordered
    ``("HH:MM", name)`` pairs (the first at 00:00); it drives the dynamic
    columns of the "Tarifzonen" sheet and the "Definition" sheet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    zones: list[tuple[str, str]] = [
        (t, name) for t, name in meta.get("zones", [])
    ]
    inner_times = [t for t, _name in zones[1:]]
    zone_names = list(dict.fromkeys(name for _t, name in zones))
    # Each zone's daily windows, e.g. "Standard: 06:00-12:00 + 16:00-18:00".
    segment_ends = [*inner_times, "24:00"]
    zone_windows = {
        name: [
            f"{zones[i][0]}-{segment_ends[i]}"
            for i in range(len(zones))
            if zones[i][1] == name
        ]
        for name in zone_names
    }

    wb = Workbook()

    # --- Sheet 1: raw readings (long: one row per reading) --------------
    raw = wb.active
    raw.title = "Rohdaten"
    raw.append(["Zeitstempel", "OBIS", "Wert (kWh)", "Einheit", "Qualitaet"])
    for r in readings:
        raw.append([_fmt_dt(r.timestamp), r.obis_code, r.value, r.unit,
                    r.quality])
    for row in raw.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "0.0000"

    # --- Sheet 2: daily end values --------------------------------------
    ende = wb.create_sheet("Tagesendwerte")
    ende.append([
        "Datum", "verwendeter Zeitstempel",
        "1.8.0 Bezug (kWh)", "2.8.0 Einspeisung (kWh)",
    ])
    for s in daily_summary:
        ende.append([
            s.day.isoformat(), _fmt_dt(s.end_timestamp),
            s.import_end, s.export_end,
        ])
    for row in ende.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = "0.0000"

    # --- Sheet 3: tariff zones (dynamic columns from the zone config) ----
    tarif = wb.create_sheet("Tarifzonen")
    tarif.append([
        "Datum",
        "Bezug 00:00 (kWh)",
        *[f"Bezug {t} (kWh)" for t in inner_times],
        "Bezug Folgetag 00:00 (kWh)",
        *[f"Verbrauch {name} (kWh)" for name in zone_names],
        "Gesamtverbrauch 00:00-24:00 (kWh)",
        "Einspeisung gesamt (kWh)",
    ])
    for s in daily_summary:
        tarif.append([
            s.day.isoformat(),
            s.import_start,
            *[s.import_switches.get(t) for t in inner_times],
            s.import_end,
            *[s.zone_consumptions.get(name) for name in zone_names],
            s.consumption_total,
            s.feedin_total,
        ])
    for row in tarif.iter_rows(min_row=2, min_col=2, max_col=tarif.max_column):
        for cell in row:
            cell.number_format = "0.0000"

    # --- Sheet 4: definitions (layout mirrors the standalone v6 script) --
    info = wb.create_sheet("Definition")
    bold = Font(bold=True)
    info["A1"] = "Definitionen"
    info["A1"].font = bold
    info["A3"] = "Export"
    info["A3"].font = bold
    info["A4"] = (
        f"Zähler: {meta.get('meter_id', '')}    "
        f"Zeitraum: {meta.get('from', '')} bis {meta.get('to', '')}"
    )
    info["A6"] = "Tagesendwert"
    info["A6"].font = bold
    info["A7"] = (
        "Der Tagesendwert eines Tages D ist der erste vorhandene kumulative "
        "Zählerstand am Folgetag D+1 um 00:00 Uhr lokaler Zeit "
        "(Europe/Berlin)."
    )
    info["A9"] = "Tarifzonen (konfiguriert, lokale Zeit des Kalendertags D)"
    info["A9"].font = bold
    row = 10
    for name in zone_names:
        info[f"A{row}"] = f"{name}: {' + '.join(zone_windows[name])}"
        row += 1
    row += 1
    info[f"A{row}"] = "Berechnung Bezug"
    info[f"A{row}"].font = bold
    row += 1
    info[f"A{row}"] = (
        "Verbrauch eines Zeitfensters = Bezug(Fensterende) - "
        "Bezug(Fensterbeginn); 24:00 entspricht 00:00 des Folgetags D+1."
    )
    row += 1
    info[f"A{row}"] = (
        "Verbrauch einer Tarifzone = Summe der Verbräuche ihrer Zeitfenster."
    )
    row += 1
    info[f"A{row}"] = "Gesamtverbrauch = Bezug(Folgetag 00:00) - Bezug(00:00)"
    row += 2
    info[f"A{row}"] = "Wichtiger Hinweis"
    info[f"A{row}"].font = bold
    row += 1
    info[f"A{row}"] = (
        "Gesamtverbrauch wird direkt aus den beiden Tagesrand-Zählerständen "
        "berechnet. Er ist damit rechnerisch identisch zur Summe der "
        "Tarifzonen-Verbräuche, sofern alle Messpunkte vorhanden sind."
    )
    row += 1
    info[f"A{row}"] = (
        "Falls für einen Tag ein benötigter Messpunkt fehlt (00:00 oder "
        "eine Umschaltzeit), bleibt die betroffene berechnete Spalte leer."
    )
    info.column_dimensions["A"].width = 140

    # --- header styling + autosize for all data sheets ------------------
    for ws in (raw, ende, tarif):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max(
                len("" if c.value is None else str(c.value))
                for c in column_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_len + 2, 40
            )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)
